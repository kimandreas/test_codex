from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from test_codex.chatgpt_excel import load_prompt_jobs
from test_codex.chatgpt_excel import read_range_as_prompt
from test_codex.chatgpt_excel import run_prompt_jobs
from test_codex.chatgpt_excel import run_chatgpt_prompt_from_workbook
from test_codex.excel_agent import apply_note_mappings, combine_excel_notes, combine_notes
from test_codex.excel_agent import load_mappings
from test_codex.excel_agent import write_combined_notes_to_workbook


def test_combine_notes_skips_empty_cells() -> None:
    notes = ["First note", None, "  ", "Second note"]

    assert combine_notes(notes) == "First note\n\nSecond note"


def test_combine_notes_uses_custom_separator() -> None:
    notes = ["A", "B", "C"]

    assert combine_notes(notes, separator=" | ") == "A | B | C"


def test_combine_excel_notes_writes_target_cell(tmp_path) -> None:
    workbook_path = tmp_path / "notes.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "Visible value ignored"
    worksheet["B1"] = "Another visible value ignored"
    worksheet["A1"].comment = Comment("First note", "Codex")
    worksheet["B1"].comment = Comment("Second note", "Codex")
    workbook.save(workbook_path)

    combined = combine_excel_notes(workbook_path, ["A1", "B1"], "C1")

    saved_workbook = load_workbook(workbook_path)
    assert combined == "First note\n\nSecond note"
    assert saved_workbook.active["C1"].value == "First note\n\nSecond note"
    assert saved_workbook.active["C1"].alignment.wrap_text is True


def test_combine_excel_notes_can_read_cell_values(tmp_path) -> None:
    workbook_path = tmp_path / "values.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "First value"
    worksheet["B1"] = "Second value"
    workbook.save(workbook_path)

    combined = combine_excel_notes(
        workbook_path,
        ["A1", "B1"],
        "C1",
        use_cell_values=True,
    )

    assert combined == "First value\n\nSecond value"


def test_write_combined_notes_to_existing_output_workbook(tmp_path) -> None:
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "output.xlsx"

    source_workbook = Workbook()
    source_worksheet = source_workbook.active
    source_worksheet.title = "Source Notes"
    source_worksheet["C4"].comment = Comment("First note", "Codex")
    source_worksheet["C9"].comment = Comment("Second note", "Codex")
    source_workbook.save(source_path)

    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = "Summary"
    output_worksheet["D36"] = "Existing value"
    output_workbook.save(output_path)

    combined = write_combined_notes_to_workbook(
        source_path,
        output_path,
        ["C4", "C9"],
        "D36",
        source_sheet_name="Source Notes",
        target_sheet_name="Summary",
    )

    saved_output = load_workbook(output_path)
    assert combined == "First note\n\nSecond note"
    assert saved_output["Summary"]["D36"].value == "First note\n\nSecond note"
    assert saved_output["Summary"]["D36"].alignment.wrap_text is True


def test_load_mappings_from_csv(tmp_path) -> None:
    mapping_path = tmp_path / "mapping.csv"
    mapping_path.write_text(
        "source_sheet,source_cells,target_sheet,target_cell\n"
        "Source Notes,\"C4, C9\",Summary,D36\n",
        encoding="utf-8",
    )

    mappings = load_mappings(mapping_path)

    assert len(mappings) == 1
    assert mappings[0].source_sheet == "Source Notes"
    assert mappings[0].source_cells == ["C4", "C9"]
    assert mappings[0].target_sheet == "Summary"
    assert mappings[0].target_cell == "D36"


def test_apply_note_mappings_writes_multiple_targets(tmp_path) -> None:
    source_path = tmp_path / "source.xlsx"
    output_path = tmp_path / "output.xlsx"

    source_workbook = Workbook()
    source_worksheet = source_workbook.active
    source_worksheet.title = "Source Notes"
    source_worksheet["C4"].comment = Comment("First note", "Codex")
    source_worksheet["C9"].comment = Comment("Second note", "Codex")
    source_worksheet["C11"].comment = Comment("Third note", "Codex")
    source_workbook.save(source_path)

    output_workbook = Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = "Summary"
    output_workbook.save(output_path)

    mapping_path = tmp_path / "mapping.csv"
    mapping_path.write_text(
        "source_sheet,source_cells,target_sheet,target_cell\n"
        "Source Notes,\"C4, C9\",Summary,D36\n"
        "Source Notes,C11,Summary,D37\n",
        encoding="utf-8",
    )

    results = apply_note_mappings(
        source_path,
        output_path,
        load_mappings(mapping_path),
    )

    saved_output = load_workbook(output_path)
    assert [combined for _, combined in results] == [
        "First note\n\nSecond note",
        "Third note",
    ]
    assert saved_output["Summary"]["D36"].value == "First note\n\nSecond note"
    assert saved_output["Summary"]["D37"].value == "Third note"


def test_read_range_as_prompt_preserves_table_shape() -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "Role"
    worksheet["B1"] = "You are a helpful reviewer."
    worksheet["A2"] = "Task"
    worksheet["B2"] = "Summarize the notes."

    assert (
        read_range_as_prompt(worksheet, "A1:B2")
        == "Role\tYou are a helpful reviewer.\nTask\tSummarize the notes."
    )


def test_run_chatgpt_prompt_from_workbook_writes_response(tmp_path) -> None:
    source_path = tmp_path / "prompt.xlsx"
    output_path = tmp_path / "response.xlsx"

    source_workbook = Workbook()
    prompt_sheet = source_workbook.active
    prompt_sheet.title = "Prompt"
    prompt_sheet["A1"] = "Please summarize:"
    prompt_sheet["A2"] = "Student participated well."
    source_workbook.save(source_path)

    output_workbook = Workbook()
    response_sheet = output_workbook.active
    response_sheet.title = "Results"
    output_workbook.save(output_path)

    def fake_response(prompt: str, model: str, instructions: str | None) -> str:
        assert prompt == "Please summarize:\nStudent participated well."
        assert model == "fake-model"
        assert instructions == "Be concise."
        return "Strong participation."

    prompt, response = run_chatgpt_prompt_from_workbook(
        source_path,
        output_path,
        prompt_range="A1:A2",
        prompt_sheet_name="Prompt",
        response_sheet_name="Results",
        response_cell="B2",
        model="fake-model",
        instructions="Be concise.",
        response_generator=fake_response,
    )

    saved_output = load_workbook(output_path)
    assert prompt == "Please summarize:\nStudent participated well."
    assert response == "Strong participation."
    assert saved_output["Results"]["B2"].value == "Strong participation."
    assert saved_output["Results"]["B2"].alignment.wrap_text is True


def test_run_chatgpt_prompt_reads_prompt_from_first_workbook(tmp_path) -> None:
    prompt_path = tmp_path / "assessment_tool.xlsx"
    response_path = tmp_path / "test.xlsx"

    prompt_workbook = Workbook()
    prompt_sheet = prompt_workbook.active
    prompt_sheet.title = "Prompt"
    prompt_sheet["A1"] = "Assess this class."
    prompt_workbook.save(prompt_path)

    response_workbook = Workbook()
    result_sheet = response_workbook.active
    result_sheet.title = "Results"
    response_workbook.save(response_path)

    def fake_response(prompt: str, model: str, instructions: str | None) -> str:
        assert prompt == "Assess this class."
        return "Assessment complete."

    run_chatgpt_prompt_from_workbook(
        prompt_path,
        response_path,
        prompt_range="A1:A1",
        prompt_sheet_name="Prompt",
        response_sheet_name="Results",
        response_cell="B2",
        response_generator=fake_response,
    )

    saved_output = load_workbook(response_path)
    assert saved_output["Results"]["B2"].value == "Assessment complete."


def test_load_prompt_jobs_from_csv(tmp_path) -> None:
    prompt_jobs_path = tmp_path / "prompt_jobs.csv"
    prompt_jobs_path.write_text(
        "prompt_sheet,prompt_range,response_sheet,response_cell\n"
        "Sheet2,A1:B8,Sheet1,B2\n",
        encoding="utf-8",
    )

    jobs = load_prompt_jobs(prompt_jobs_path)

    assert len(jobs) == 1
    assert jobs[0].prompt_sheet == "Sheet2"
    assert jobs[0].prompt_range == "A1:B8"
    assert jobs[0].response_sheet == "Sheet1"
    assert jobs[0].response_cell == "B2"


def test_run_prompt_jobs_writes_multiple_responses(tmp_path) -> None:
    workbook_path = tmp_path / "assessment_tool.xlsx"

    workbook = Workbook()
    prompt_sheet = workbook.active
    prompt_sheet.title = "Sheet2"
    prompt_sheet["A1"] = "Prompt one"
    prompt_sheet["A10"] = "Prompt two"
    workbook.create_sheet("Sheet1")
    workbook.save(workbook_path)

    prompt_jobs_path = tmp_path / "prompt_jobs.csv"
    prompt_jobs_path.write_text(
        "prompt_sheet,prompt_range,response_sheet,response_cell\n"
        "Sheet2,A1:A1,Sheet1,B2\n"
        "Sheet2,A10:A10,Sheet1,B7\n",
        encoding="utf-8",
    )

    def fake_response(prompt: str, model: str, instructions: str | None) -> str:
        return f"response for {prompt}"

    results = run_prompt_jobs(
        workbook_path,
        workbook_path,
        load_prompt_jobs(prompt_jobs_path),
        response_generator=fake_response,
    )

    saved_workbook = load_workbook(workbook_path)
    assert [response for _, response in results] == [
        "response for Prompt one",
        "response for Prompt two",
    ]
    assert saved_workbook["Sheet1"]["B2"].value == "response for Prompt one"
    assert saved_workbook["Sheet1"]["B7"].value == "response for Prompt two"
