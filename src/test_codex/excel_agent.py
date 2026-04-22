"""Helpers for combining notes from Excel workbooks."""

from __future__ import annotations

import csv
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass(frozen=True)
class NoteMapping:
    """A single note transfer from source cells to one target cell."""

    source_cells: list[str]
    target_sheet: str
    target_cell: str
    source_sheet: str | None = None


def clean_note(value: Any) -> str:
    """Convert note text to a clean string."""
    if value is None:
        return ""
    return str(value).strip()


def combine_notes(notes: list[Any], separator: str = "\n\n") -> str:
    """Combine non-empty notes with the requested separator."""
    cleaned_notes = [clean_note(note) for note in notes]
    return separator.join(note for note in cleaned_notes if note)


def read_cell_note(worksheet: Any, cell_address: str) -> str:
    """Read text from a worksheet cell's attached Excel note/comment."""
    cell = worksheet[cell_address]
    return clean_note(cell.comment.text if cell.comment else None)


def read_cell_value(worksheet: Any, cell_address: str) -> str:
    """Read text from a worksheet cell value."""
    cell = worksheet[cell_address]
    return clean_note(cell.value)


def get_worksheet(workbook: Any, sheet_name: str | None) -> Any:
    """Return the requested worksheet, or the active worksheet when omitted."""
    return workbook[sheet_name] if sheet_name else workbook.active


def write_combined_note(worksheet: Any, target_cell: str, combined_note: str) -> None:
    """Write combined note text to a target cell and enable wrapping."""
    target = worksheet[target_cell]
    target.value = combined_note

    try:
        from openpyxl.styles import Alignment
    except ImportError:
        Alignment = None

    if Alignment:
        alignment = copy(target.alignment)
        alignment.wrap_text = True
        target.alignment = alignment


def combine_excel_notes(
    workbook_path: str | Path,
    source_cells: list[str],
    target_cell: str,
    *,
    sheet_name: str | None = None,
    output_path: str | Path | None = None,
    separator: str = "\n\n",
    use_cell_values: bool = False,
) -> str:
    """Read Excel notes from source cells, write the combined result, and save."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        message = "openpyxl is required. Install it with: python -m pip install -e '.[test]'"
        raise RuntimeError(message) from exc

    output_file = Path(output_path) if output_path else workbook_path

    combined_note = write_combined_notes_to_workbook(
        source_workbook_path=workbook_path,
        output_workbook_path=output_file,
        source_cells=source_cells,
        target_cell=target_cell,
        source_sheet_name=sheet_name,
        target_sheet_name=sheet_name,
        separator=separator,
        use_cell_values=use_cell_values,
    )
    return combined_note


def write_combined_notes_to_workbook(
    source_workbook_path: str | Path,
    output_workbook_path: str | Path,
    source_cells: list[str],
    target_cell: str,
    *,
    source_sheet_name: str | None = None,
    target_sheet_name: str | None = None,
    separator: str = "\n\n",
    use_cell_values: bool = False,
) -> str:
    """Read notes from one workbook and write them to an existing output workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        message = "openpyxl is required. Install it with: python -m pip install -e '.[test]'"
        raise RuntimeError(message) from exc

    source_workbook = load_workbook(Path(source_workbook_path))
    output_workbook = load_workbook(Path(output_workbook_path))

    source_worksheet = get_worksheet(source_workbook, source_sheet_name)
    target_worksheet = get_worksheet(output_workbook, target_sheet_name)

    reader = read_cell_value if use_cell_values else read_cell_note
    notes = [reader(source_worksheet, source_cell) for source_cell in source_cells]
    combined_note = combine_notes(notes, separator=separator)
    write_combined_note(target_worksheet, target_cell, combined_note)

    output_workbook.save(Path(output_workbook_path))
    return combined_note


def parse_cell_list(value: str) -> list[str]:
    """Parse source cells from comma-separated or whitespace-separated text."""
    normalized = value.replace(",", " ")
    return [cell.strip() for cell in normalized.split() if cell.strip()]


def load_mappings(mapping_csv_path: str | Path) -> list[NoteMapping]:
    """Load note transfer mappings from a CSV file."""
    mappings: list[NoteMapping] = []
    with Path(mapping_csv_path).open(newline="", encoding="utf-8-sig") as mapping_file:
        reader = csv.DictReader(mapping_file)
        required_fields = {"source_cells", "target_sheet", "target_cell"}
        missing_fields = required_fields - set(reader.fieldnames or [])
        if missing_fields:
            missing = ", ".join(sorted(missing_fields))
            raise ValueError(f"Mapping CSV is missing required column(s): {missing}")

        for row_number, row in enumerate(reader, start=2):
            source_cells = parse_cell_list(row.get("source_cells", ""))
            target_sheet = clean_note(row.get("target_sheet"))
            target_cell = clean_note(row.get("target_cell"))
            source_sheet = clean_note(row.get("source_sheet")) or None

            if not source_cells or not target_sheet or not target_cell:
                raise ValueError(
                    "Mapping CSV row "
                    f"{row_number} must include source_cells, target_sheet, and target_cell"
                )

            mappings.append(
                NoteMapping(
                    source_cells=source_cells,
                    source_sheet=source_sheet,
                    target_sheet=target_sheet,
                    target_cell=target_cell,
                )
            )

    return mappings


def apply_note_mappings(
    source_workbook_path: str | Path,
    output_workbook_path: str | Path,
    mappings: list[NoteMapping],
    *,
    default_source_sheet_name: str | None = None,
    separator: str = "\n\n",
    use_cell_values: bool = False,
) -> list[tuple[NoteMapping, str]]:
    """Apply multiple note mappings and save the output workbook once."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        message = "openpyxl is required. Install it with: python -m pip install -e '.[test]'"
        raise RuntimeError(message) from exc

    source_workbook = load_workbook(Path(source_workbook_path))
    output_workbook = load_workbook(Path(output_workbook_path))
    reader = read_cell_value if use_cell_values else read_cell_note
    results: list[tuple[NoteMapping, str]] = []

    for mapping in mappings:
        source_sheet_name = mapping.source_sheet or default_source_sheet_name
        source_worksheet = get_worksheet(source_workbook, source_sheet_name)
        target_worksheet = get_worksheet(output_workbook, mapping.target_sheet)
        notes = [reader(source_worksheet, source_cell) for source_cell in mapping.source_cells]
        combined_note = combine_notes(notes, separator=separator)
        write_combined_note(target_worksheet, mapping.target_cell, combined_note)
        results.append((mapping, combined_note))

    output_workbook.save(Path(output_workbook_path))
    return results
