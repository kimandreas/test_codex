"""Read Excel prompt ranges, call OpenAI, and write responses back to Excel."""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable

from test_codex.excel_agent import clean_note, get_worksheet, write_combined_note


ResponseGenerator = Callable[[str, str, str | None], str]


@dataclass(frozen=True)
class PromptJob:
    """A single prompt range to response cell job."""

    prompt_sheet: str
    prompt_range: str
    response_sheet: str
    response_cell: str


def read_range_as_prompt(worksheet: Any, cell_range: str) -> str:
    """Read a cell range as tab-separated prompt text."""
    lines: list[str] = []
    for row in worksheet[cell_range]:
        values = [clean_note(cell.value) for cell in row]
        while values and not values[-1]:
            values.pop()
        if values:
            lines.append("\t".join(values))
    return "\n".join(lines).strip()


def generate_chatgpt_response(
    prompt: str,
    model: str,
    instructions: str | None = None,
) -> str:
    """Send prompt text to OpenAI's Responses API and return text output."""
    try:
        from openai import OpenAI
    except ImportError as exc:
        message = "openai is required. Install it with: python -m pip install -e '.[test]'"
        raise RuntimeError(message) from exc

    if not os.environ.get("OPENAI_API_KEY"):
        raise RuntimeError("OPENAI_API_KEY is required to call the OpenAI API.")

    client = OpenAI()
    response = client.responses.create(
        model=model,
        input=prompt,
        instructions=instructions,
    )
    return response.output_text.strip()


def run_chatgpt_prompt_from_workbook(
    source_workbook_path: str | Path,
    output_workbook_path: str | Path,
    *,
    prompt_range: str,
    response_sheet_name: str,
    response_cell: str,
    prompt_sheet_name: str | None = None,
    model: str = "gpt-4.1-mini",
    instructions: str | None = None,
    prompt_workbook_path: str | Path | None = None,
    response_generator: ResponseGenerator = generate_chatgpt_response,
) -> tuple[str, str]:
    """Read a prompt range, generate a response, write it, and save the output."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        message = "openpyxl is required. Install it with: python -m pip install -e '.[test]'"
        raise RuntimeError(message) from exc

    prompt_workbook_file = Path(prompt_workbook_path or source_workbook_path)
    source_workbook = load_workbook(prompt_workbook_file, data_only=True)
    output_workbook = load_workbook(Path(output_workbook_path))

    prompt_worksheet = get_worksheet(source_workbook, prompt_sheet_name)
    response_worksheet = get_worksheet(output_workbook, response_sheet_name)

    prompt = read_range_as_prompt(prompt_worksheet, prompt_range)
    if not prompt:
        raise ValueError(f"Prompt range {prompt_range} did not contain any text.")

    response_text = response_generator(prompt, model, instructions)
    write_combined_note(response_worksheet, response_cell, response_text)
    output_workbook.save(Path(output_workbook_path))
    return prompt, response_text


def load_prompt_jobs(prompt_jobs_csv_path: str | Path) -> list[PromptJob]:
    """Load prompt jobs from a CSV file."""
    jobs: list[PromptJob] = []
    with Path(prompt_jobs_csv_path).open(newline="", encoding="utf-8-sig") as jobs_file:
        reader = csv.DictReader(jobs_file)
        required_fields = {
            "prompt_sheet",
            "prompt_range",
            "response_sheet",
            "response_cell",
        }
        missing_fields = required_fields - set(reader.fieldnames or [])
        if missing_fields:
            missing = ", ".join(sorted(missing_fields))
            raise ValueError(f"Prompt jobs CSV is missing required column(s): {missing}")

        for row_number, row in enumerate(reader, start=2):
            job = PromptJob(
                prompt_sheet=clean_note(row.get("prompt_sheet")),
                prompt_range=clean_note(row.get("prompt_range")),
                response_sheet=clean_note(row.get("response_sheet")),
                response_cell=clean_note(row.get("response_cell")),
            )
            if not all(
                [job.prompt_sheet, job.prompt_range, job.response_sheet, job.response_cell]
            ):
                raise ValueError(
                    "Prompt jobs CSV row "
                    f"{row_number} must include prompt_sheet, prompt_range, "
                    "response_sheet, and response_cell"
                )
            jobs.append(job)

    return jobs


def run_prompt_jobs(
    prompt_workbook_path: str | Path,
    response_workbook_path: str | Path,
    jobs: list[PromptJob],
    *,
    model: str = "gpt-4.1-mini",
    instructions: str | None = None,
    response_generator: ResponseGenerator = generate_chatgpt_response,
) -> list[tuple[PromptJob, str]]:
    """Run multiple prompt jobs and save the response workbook after each response."""
    results: list[tuple[PromptJob, str]] = []
    for job in jobs:
        _, response_text = run_chatgpt_prompt_from_workbook(
            prompt_workbook_path,
            response_workbook_path,
            prompt_range=job.prompt_range,
            prompt_sheet_name=job.prompt_sheet,
            response_sheet_name=job.response_sheet,
            response_cell=job.response_cell,
            model=model,
            instructions=instructions,
            response_generator=response_generator,
        )
        results.append((job, response_text))
    return results
